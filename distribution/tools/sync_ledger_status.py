"""Make the activation-key workbook enforceable: sync Status to D1 daily.

The workbook (``distribution/ledger/activation_keys_*.xlsx``) is where codes
are marked Sold / Activated / Replaced / Void — but until this sync existed it
was pure bookkeeping. Now a scheduled run reads the Status column and sets the
``disabled`` flag in the live D1 database, and the Worker refuses a disabled
code for activation, re-activation and download.

Only Unused stays live — the owner's explicit rule (2026-08-10):

  Unused     active   — the shelf. Reverting any code to Unused in the sheet
                        re-enables it on the next run.
  Sold       48-HOUR WINDOW — the buyer is told they have 48 hours to
                        install. A Sold code is retired at the first run at
                        least ``--sold-window-days`` (default 2) days after
                        its Sold Date — and even then only if it is STILL
                        UNCLAIMED in the database, so a buyer who activated
                        while the sheet went un-updated is never punished for
                        the bookkeeping. A Sold row with no readable Sold
                        Date is left active and warned about.
  Activated  RETIRED  — the buyer's install keeps running (the licence
                        validates offline), but the code itself stops
                        answering. A reinstall after a wiped disk therefore
                        needs a replacement code — accepted support cost.
  Replaced   RETIRED  — a replacement was issued; without this the original
                        device could re-activate forever alongside it.
  Void       RETIRED  — minted but cancelled or leaked.

What retiring does NOT do: reach into an installed, activated copy. The app
validates its cached licence offline by design; this controls the code, not
the customer's machine.

Run manually::

    .venv\\Scripts\\python.exe -m distribution.tools.sync_ledger_status            # apply
    .venv\\Scripts\\python.exe -m distribution.tools.sync_ledger_status --dry-run  # show SQL only

The scheduled task (see ``register_ledger_sync_task.ps1`` beside this file)
runs the apply form daily at 10:00 local time.
"""

from __future__ import annotations

import argparse
import datetime as _datetime
import subprocess
import sys
from pathlib import Path
from typing import NamedTuple

from openpyxl import load_workbook

from distribution.licensing.codes import normalize_code

LEDGER_DIR = Path(__file__).resolve().parents[1] / "ledger"
ACTIVATION_SERVER_DIR = Path(__file__).resolve().parents[1] / "activation-server"
SHEET_NAME = "Activation Keys"
D1_DATABASE = "pitwall-licenses"

# Statuses whose codes are retired outright. Sold is absent because it is
# governed by the 48-hour window, not because it survives — see the module
# docstring. Only Unused stays live.
DEFAULT_DISABLING_STATUSES = frozenset({"Activated", "Replaced", "Void"})
KNOWN_STATUSES = frozenset({"Unused", "Sold", "Activated", "Replaced", "Void"})


def newest_workbook(ledger_dir: Path = LEDGER_DIR) -> Path:
    candidates = sorted(ledger_dir.glob("activation_keys_*.xlsx"))
    if not candidates:
        raise FileNotFoundError(
            f"No activation_keys_*.xlsx found in {ledger_dir}. "
            "Generate one with distribution.tools.generate_codes first."
        )
    return candidates[-1]


class LedgerRow(NamedTuple):
    status: str
    sold_date: _datetime.date | None


def _parse_sold_date(value: object) -> _datetime.date | None:
    """Sold Date as a date, or None when absent or unreadable.

    Excel-native dates arrive as datetime; hand-typed values are accepted in
    ISO form only. Anything else is None — an unreadable date must never
    START a countdown that ends in a retired code.
    """
    if isinstance(value, _datetime.datetime):
        return value.date()
    if isinstance(value, _datetime.date):
        return value
    if isinstance(value, str) and value.strip():
        try:
            return _datetime.date.fromisoformat(value.strip())
        except ValueError:
            return None
    return None


def read_statuses(workbook_path: Path) -> dict[str, LedgerRow]:
    """Map canonical code -> (status, sold date), with loud failures.

    A typo'd code or unknown status raises rather than being skipped: silently
    ignoring a row would leave a code the ledger says is dead still working,
    which is the exact failure this tool exists to remove.
    """
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"{workbook_path.name} has no '{SHEET_NAME}' sheet.")
    sheet = workbook[SHEET_NAME]
    statuses: dict[str, LedgerRow] = {}
    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
        raw_code, raw_status, _buyer, raw_sold_date = (row + (None,) * 4)[:4]
        if raw_code is None or str(raw_code).strip() == "":
            continue
        code = normalize_code(str(raw_code))
        if code is None:
            raise ValueError(f"Row with unreadable code: {raw_code!r}")
        status = str(raw_status or "Unused").strip().title()
        if status not in KNOWN_STATUSES:
            raise ValueError(
                f"{code}: unknown status {raw_status!r}. "
                f"Expected one of: {', '.join(sorted(KNOWN_STATUSES))}."
            )
        if code in statuses:
            raise ValueError(f"{code} appears twice in the workbook.")
        statuses[code] = LedgerRow(status, _parse_sold_date(raw_sold_date))
    if not statuses:
        raise ValueError(f"{workbook_path.name} contains no codes.")
    return statuses


def sold_window_expired(
    row: LedgerRow,
    today: _datetime.date,
    window_days: int,
) -> bool:
    """Whether a Sold code's activation window has run out.

    Measured in whole days from the workbook's Sold Date, so with the default
    of 2 the code dies at the first run at least two days after the sale —
    every buyer gets their full promised 48 hours regardless of when the run
    happens. No Sold Date means no countdown.
    """
    if row.status != "Sold" or row.sold_date is None:
        return False
    return (today - row.sold_date).days >= window_days


def build_sql(
    statuses: dict[str, LedgerRow],
    disabling: frozenset[str] = DEFAULT_DISABLING_STATUSES,
    *,
    today: _datetime.date | None = None,
    sold_window_days: int = 2,
) -> str:
    """One idempotent UPDATE per code; the workbook is authoritative.

    Enabling is as deliberate as disabling: moving a code back to Unused in
    the sheet re-enables it on the next run, so a mis-click is recoverable by
    fixing the sheet rather than by editing the database.

    A Sold code past its window carries ``AND claimed = 0``: if the buyer
    activated but the sheet was never updated, the code is de facto Activated
    and must keep answering same-device re-activation. D1 evaluates the guard
    atomically, so the sync never needs to read state first.
    """
    # Local date on purpose: the sheet is maintained in local time and the
    # scheduled run fires at 10:00 local, so the window must count local days.
    today = today or _datetime.date.today()  # noqa: DTZ011
    header = (
        f"-- Disabling statuses: {', '.join(sorted(disabling))}; "
        f"Sold window: {sold_window_days} day(s)"
    )
    lines = [
        "-- Generated by distribution.tools.sync_ledger_status; do not edit.",
        header,
    ]
    for code, row in sorted(statuses.items()):
        if row.status in disabling:
            lines.append(
                "UPDATE codes SET disabled = 1, "
                f"disabled_reason = 'ledger status: {row.status}' "
                f"WHERE code_id = '{code}';"
            )
        elif sold_window_expired(row, today, sold_window_days):
            lines.append(
                "UPDATE codes SET disabled = 1, "
                "disabled_reason = 'ledger status: Sold "
                f"(activation window expired {row.sold_date.isoformat()} + {sold_window_days}d)' "
                f"WHERE code_id = '{code}' AND claimed = 0;"
            )
        else:
            lines.append(
                "UPDATE codes SET disabled = 0, disabled_reason = NULL "
                f"WHERE code_id = '{code}';"
            )
    return "\n".join(lines) + "\n"


def run_wrangler(sql_path: Path, *, remote: bool) -> int:
    command = [
        "npx",
        "wrangler",
        "d1",
        "execute",
        D1_DATABASE,
        "--remote" if remote else "--local",
        "--file",
        str(sql_path),
        "-y",
    ]
    completed = subprocess.run(
        command,
        cwd=ACTIVATION_SERVER_DIR,
        shell=sys.platform == "win32",  # npx is npx.cmd on Windows
        check=False,  # the caller reports a non-zero exit; raising adds nothing
    )
    return completed.returncode


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook", type=Path, default=None,
        help="explicit workbook path (default: newest in distribution/ledger)",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="print the SQL and what would change, execute nothing",
    )
    parser.add_argument(
        "--local", action="store_true",
        help="apply to the local wrangler dev database instead of production",
    )
    parser.add_argument(
        "--sold-window-days", type=int, default=2,
        help="days a Sold code stays activatable after its Sold Date (default 2 = the promised 48 hours)",
    )
    args = parser.parse_args(argv)

    disabling = set(DEFAULT_DISABLING_STATUSES)

    workbook_path = args.workbook or newest_workbook()
    statuses = read_statuses(workbook_path)
    today = _datetime.date.today()  # noqa: DTZ011 - local days, see build_sql
    sql = build_sql(
        statuses,
        frozenset(disabling),
        today=today,
        sold_window_days=args.sold_window_days,
    )

    retired = sorted(c for c, r in statuses.items() if r.status in disabling)
    expired = sorted(
        c for c, r in statuses.items()
        if sold_window_expired(r, today, args.sold_window_days)
    )
    undated = sorted(
        c for c, r in statuses.items() if r.status == "Sold" and r.sold_date is None
    )
    stamp = _datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # noqa: DTZ005 - local log line
    print(f"[{stamp}] {workbook_path.name}: {len(statuses)} codes; "
          f"{len(retired)} retired by statuses {sorted(disabling)}; "
          f"{len(expired)} Sold past the {args.sold_window_days}-day window")
    for code in retired:
        print(f"  retired: {code} ({statuses[code].status})")
    for code in expired:
        print(f"  window expired (if still unclaimed): {code} "
              f"(sold {statuses[code].sold_date})")
    for code in undated:
        print(f"  WARNING: {code} is Sold with no readable Sold Date — "
              "left active, no countdown is running")

    if args.dry_run:
        print("\n--dry-run: SQL below was NOT executed.\n")
        print(sql)
        return 0

    sql_path = LEDGER_DIR / "sync_ledger_status.generated.sql"
    sql_path.write_text(sql, encoding="utf-8")
    returncode = run_wrangler(sql_path, remote=not args.local)
    if returncode != 0:
        print(f"wrangler exited {returncode}; the database was NOT fully synced.")
    return returncode


if __name__ == "__main__":
    sys.exit(main())
