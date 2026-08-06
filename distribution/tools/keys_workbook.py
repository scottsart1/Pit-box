"""Build the activation-key tracker workbook.

This is the sheet you actually work from when someone pays: find the first
Unused code, mark it Sold, record who it went to, send it.

It is written by `generate_codes.py` in the same run that mints the codes and
the D1 seed, so the three cannot drift apart. It lands in `distribution/ledger/`
which is gitignored — it lists every code you can sell, so it is as sensitive
as the ledger JSON. Never commit it, never email it, never put it in a shared
drive.

The workbook is a record, not a mechanism: nothing reads it back. Enforcement
of one-time use lives in the activation server (an atomic claim in D1), and
enforcement of one-computer-only lives in the app (the device hash re-checked
at every launch). This exists so you know what you have sold and to whom.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"

# Lifecycle. Deliberately small: every extra state is one more thing to get
# wrong at 1am after a sale.
STATUSES = ("Unused", "Sold", "Activated", "Replaced", "Void")

HEADERS = (
    ("Activation Code", 26),
    ("Status", 13),
    ("Buyer Email", 30),
    ("Sold Date", 13),
    ("Activated Date", 15),
    ("Device (short)", 18),
    ("Notes", 40),
)

# Which columns you type into. Shaded yellow so the generated column is
# visually distinct from the ones you maintain.
EDITABLE_COLUMNS = ("B", "C", "D", "E", "F", "G")

INK = "1F2937"
ACCENT = "C8102E"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
EDIT_FILL = PatternFill("solid", fgColor="FFF9DB")
THIN = Side(style="thin", color="D5DAE0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _title(sheet, text: str, *, row: int = 1, size: int = 14) -> None:
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT, size=size, bold=True, color=INK)


def _para(sheet, row: int, text: str, *, bold: bool = False, color: str = INK) -> int:
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT, size=10, bold=bold, color=color)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


def _build_codes_sheet(sheet, codes: list[str]) -> None:
    sheet.title = "Activation Keys"

    for index, (header, width) in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 22

    for offset, code in enumerate(codes):
        row = offset + 2
        code_cell = sheet.cell(row=row, column=1, value=code)
        # Monospace so a mistyped code is visible when read back to a buyer.
        code_cell.font = Font(name="Consolas", size=10, color=INK)
        code_cell.border = BORDER

        status = sheet.cell(row=row, column=2, value="Unused")
        status.font = Font(name=FONT, size=10, color=INK)

        for column in range(2, len(HEADERS) + 1):
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.fill = EDIT_FILL
            cell.border = BORDER
        for column in (4, 5):
            sheet.cell(row=row, column=column).number_format = "yyyy-mm-dd"

    # An empty batch would give "B2:B1", which is not a range: openpyxl raises
    # and the whole mint fails. Clamping keeps the sheet valid and empty.
    last = max(len(codes) + 1, 2)

    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUSES) + '"',
        allow_blank=False,
        showDropDown=False,  # False means "show the dropdown arrow" in OOXML
    )
    validation.error = "Pick one of: " + ", ".join(STATUSES)
    validation.errorTitle = "Not a valid status"
    sheet.add_data_validation(validation)
    validation.add(f"B2:B{last}")

    # Colour the status column so the sellable stock is findable at a glance.
    for value, colour in (
        ("Unused", "1B7F4B"),
        ("Sold", "B26B00"),
        ("Activated", "1F2937"),
        ("Replaced", "6B7280"),
        ("Void", ACCENT),
    ):
        sheet.conditional_formatting.add(
            f"B2:B{last}",
            CellIsRule(
                operator="equal",
                formula=[f'"{value}"'],
                font=Font(name=FONT, size=10, bold=value in {"Unused", "Void"}, color=colour),
            ),
        )

    sheet.freeze_panes = "B2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last}"


def _build_instructions_sheet(sheet, batch_label: str) -> None:
    sheet.title = "How to use this"
    sheet.column_dimensions["A"].width = 108
    sheet.sheet_view.showGridLines = False

    _title(sheet, "Pit Wall — activation key tracker", size=16)
    row = 3
    row = _para(sheet, row, f"Batch: {batch_label}", bold=True)
    row += 1

    row = _para(sheet, row, "KEEP THIS FILE PRIVATE", bold=True, color=ACCENT)
    row = _para(
        sheet, row,
        "Every code below is worth $20 to whoever has it. Do not commit it to "
        "git, email it, or put it in a shared drive. It sits in "
        "distribution/ledger/, which is gitignored for this reason.",
    )
    row += 1

    row = _para(sheet, row, "When someone pays you on Venmo", bold=True)
    for step in (
        "1. Open the 'Activation Keys' tab and filter Status = Unused.",
        "2. Take the first code. Set its Status to Sold.",
        (
            "3. Put the buyer's email (from the Venmo payment note) in Buyer "
            "Email, and today's date in Sold Date."
        ),
        (
            "4. Email them the code. They enter it on the website to download, "
            "and again in the app the first time it runs."
        ),
        (
            "5. When they activate, set Status to Activated and fill Activated "
            "Date. You can read the short device hash off the activation server "
            "if you want it, but it is optional — it is only there to help you "
            "recognise a repeat machine."
        ),
    ):
        row = _para(sheet, row, step)
    row += 1

    row = _para(sheet, row, "Example of a filled-in row", bold=True)
    row = _para(
        sheet, row,
        "PITW-4K2QM-7XRB9-WEHTZ  |  Sold  |  buyer@example.com  |  2026-08-09  "
        "|  2026-08-09  |  3f8a…c1d2  |  Paid via Venmo, sent 9 Aug",
    )
    row += 1

    row = _para(sheet, row, "What the yellow columns mean", bold=True)
    row = _para(
        sheet, row,
        "Yellow cells are yours to fill in. The Activation Code column is "
        "generated and must never be edited — changing a character makes the "
        "code invalid, because the signature covers it.",
    )
    row += 1

    row = _para(sheet, row, "What this file does and does not do", bold=True)
    row = _para(
        sheet, row,
        "It is a record, not a lock. Nothing reads it back. One-time use is "
        "enforced by the activation server, which claims a code atomically the "
        "first time it is used — a second person entering the same code is "
        "refused even if you forgot to mark it Sold here.",
    )
    row = _para(
        sheet, row,
        "One-computer-only is enforced by the app itself. The licence is tied "
        "to a hash of the machine it activated on and re-checked at every "
        "launch, so copying an installed copy to a second computer shows a "
        "message saying the activation is for one computer only. It does not "
        "silently ask for another code.",
    )
    row += 1

    row = _para(sheet, row, "If a buyer changes computer", bold=True)
    row = _para(
        sheet, row,
        "Mark their original code Replaced, note why, and issue them a fresh "
        "Unused code. That keeps the count honest: Replaced rows are excluded "
        "from revenue on the Summary tab.",
    )


def _build_summary_sheet(sheet, count: int) -> None:
    sheet.title = "Summary"
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 16
    sheet.sheet_view.showGridLines = False

    _title(sheet, "Batch summary", size=14)

    last = max(count + 1, 2)  # see _build_codes_sheet: "B2:B1" is not a range
    codes = "'Activation Keys'"

    sheet["A3"] = "Price per licence"
    sheet["B3"] = 20
    sheet["B3"].number_format = "$#,##0"
    # Blue marks a hardcoded input: change it here and the revenue line follows.
    sheet["B3"].font = Font(name=FONT, size=10, color="0000FF", bold=True)
    sheet["A3"].font = Font(name=FONT, size=10, bold=True, color=INK)

    rows: list[tuple[str, str]] = [
        ("Codes in this batch", f"=COUNTA({codes}!$A$2:$A${last})"),
        ("Unused", f'=COUNTIF({codes}!$B$2:$B${last},"Unused")'),
        ("Sold, not yet activated", f'=COUNTIF({codes}!$B$2:$B${last},"Sold")'),
        ("Activated", f'=COUNTIF({codes}!$B$2:$B${last},"Activated")'),
        ("Replaced", f'=COUNTIF({codes}!$B$2:$B${last},"Replaced")'),
        ("Void", f'=COUNTIF({codes}!$B$2:$B${last},"Void")'),
    ]
    row = 5
    for label, formula in rows:
        sheet.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, color=INK)
        cell = sheet.cell(row=row, column=2, value=formula)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.number_format = "#,##0"
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Revenue (sold + activated)").font = Font(
        name=FONT, size=11, bold=True, color=INK
    )
    # Replaced codes are excluded on purpose: a replacement was not a second sale.
    revenue = sheet.cell(
        row=row, column=2,
        value=(
            f'=($B$3)*(COUNTIF({codes}!$B$2:$B${last},"Sold")'
            f'+COUNTIF({codes}!$B$2:$B${last},"Activated"))'
        ),
    )
    revenue.font = Font(name=FONT, size=11, bold=True, color=INK)
    revenue.number_format = "$#,##0;($#,##0);-"

    row += 2
    note = sheet.cell(
        row=row, column=1,
        value=(
            "Revenue counts Sold and Activated only. Replaced rows are "
            "excluded because a replacement code is not a second sale."
        ),
    )
    note.font = Font(name=FONT, size=9, italic=True, color="6B7280")
    note.alignment = Alignment(wrap_text=True, vertical="top")


def build_workbook(codes: list[str], destination: Path, batch_label: str) -> Path:
    """Write the tracker for one minted batch."""
    workbook = Workbook()
    _build_instructions_sheet(workbook.active, batch_label)
    _build_codes_sheet(workbook.create_sheet(), codes)
    _build_summary_sheet(workbook.create_sheet(), len(codes))

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def build_from_records(records: list[dict[str, Any]], destination: Path, batch_label: str) -> Path:
    return build_workbook([str(r["code"]) for r in records], destination, batch_label)


__all__ = ["STATUSES", "build_from_records", "build_workbook"]
