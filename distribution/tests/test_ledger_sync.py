"""The daily ledger sync: workbook Status -> D1 `disabled`, safely.

The workbook used to be pure bookkeeping. These tests pin the policy that
makes it enforceable without hurting buyers: Replaced and Void retire a code,
Sold and Activated do not (a paying buyer must still be able to activate, and
an activated buyer must still be able to re-activate after a disk death), and
the sheet is authoritative in both directions so a mis-click is undone by
fixing the sheet.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl import load_workbook

from distribution.tools.keys_workbook import build_workbook
from distribution.tools.sync_ledger_status import (
    DEFAULT_DISABLING_STATUSES,
    LedgerRow,
    build_sql,
    newest_workbook,
    read_statuses,
    sold_window_expired,
)

CODES = [
    "PITW-AAAAA-AAAAA-AAAAA",
    "PITW-BBBBB-BBBBB-BBBBB",
    "PITW-CCCCC-CCCCC-CCCCC",
    "PITW-DDDDD-DDDDD-DDDDD",
    "PITW-EEEEE-EEEEE-EEEEE",
]


def _workbook_with_statuses(tmp_path: Path, statuses: list[str]) -> Path:
    path = build_workbook(CODES, tmp_path / "activation_keys_test.xlsx", "test")
    workbook = load_workbook(path)
    sheet = workbook["Activation Keys"]
    for offset, status in enumerate(statuses):
        sheet.cell(row=2 + offset, column=2, value=status)
    workbook.save(path)
    return path


def test_reads_every_code_with_its_status(tmp_path):
    path = _workbook_with_statuses(
        tmp_path, ["Unused", "Sold", "Activated", "Replaced", "Void"]
    )
    statuses = {code: row.status for code, row in read_statuses(path).items()}
    assert statuses == {
        "PITW-AAAAA-AAAAA-AAAAA": "Unused",
        "PITW-BBBBB-BBBBB-BBBBB": "Sold",
        "PITW-CCCCC-CCCCC-CCCCC": "Activated",
        "PITW-DDDDD-DDDDD-DDDDD": "Replaced",
        "PITW-EEEEE-EEEEE-EEEEE": "Void",
    }


def test_default_policy_retires_replaced_and_void_only(tmp_path):
    """Sold and Activated stay live: the buyer paid for exactly that."""
    path = _workbook_with_statuses(
        tmp_path, ["Unused", "Sold", "Activated", "Replaced", "Void"]
    )
    sql = build_sql(read_statuses(path))
    assert (
        "UPDATE codes SET disabled = 0, disabled_reason = NULL "
        "WHERE code_id = 'PITW-BBBBB-BBBBB-BBBBB';" in sql
    ), "a Sold code must remain activatable"
    assert (
        "UPDATE codes SET disabled = 0, disabled_reason = NULL "
        "WHERE code_id = 'PITW-CCCCC-CCCCC-CCCCC';" in sql
    ), "an Activated code must keep answering same-device re-activation"
    assert (
        "UPDATE codes SET disabled = 1, disabled_reason = 'ledger status: Replaced' "
        "WHERE code_id = 'PITW-DDDDD-DDDDD-DDDDD';" in sql
    )
    assert (
        "UPDATE codes SET disabled = 1, disabled_reason = 'ledger status: Void' "
        "WHERE code_id = 'PITW-EEEEE-EEEEE-EEEEE';" in sql
    )


def test_reverting_the_sheet_re_enables_the_code(tmp_path):
    """The sheet is authoritative both ways, so a mis-click is recoverable."""
    path = _workbook_with_statuses(tmp_path, ["Void", "Unused", "Unused", "Unused", "Unused"])
    first = build_sql(read_statuses(path))
    assert (
        "UPDATE codes SET disabled = 1, disabled_reason = 'ledger status: Void' "
        "WHERE code_id = 'PITW-AAAAA-AAAAA-AAAAA';" in first
    )

    reverted = _workbook_with_statuses(
        tmp_path, ["Unused", "Unused", "Unused", "Unused", "Unused"]
    )
    second = build_sql(read_statuses(reverted))
    assert (
        "UPDATE codes SET disabled = 0, disabled_reason = NULL "
        "WHERE code_id = 'PITW-AAAAA-AAAAA-AAAAA';" in second
    )


def test_widened_policy_is_opt_in_only(tmp_path):
    path = _workbook_with_statuses(
        tmp_path, ["Unused", "Sold", "Activated", "Replaced", "Void"]
    )
    widened = frozenset(DEFAULT_DISABLING_STATUSES | {"Sold", "Activated"})
    sql = build_sql(read_statuses(path), widened)
    assert (
        "disabled = 1, disabled_reason = 'ledger status: Sold' "
        "WHERE code_id = 'PITW-BBBBB-BBBBB-BBBBB'" in sql
    )


def test_unknown_status_fails_loudly_instead_of_being_skipped(tmp_path):
    path = _workbook_with_statuses(tmp_path, ["Unused", "Refunded"])
    with pytest.raises(ValueError, match="unknown status"):
        read_statuses(path)


def test_lowercase_status_from_hand_editing_still_counts(tmp_path):
    path = _workbook_with_statuses(tmp_path, ["void"])
    assert read_statuses(path)["PITW-AAAAA-AAAAA-AAAAA"].status == "Void"


def test_sold_codes_get_the_promised_48_hours_then_die_if_unclaimed(tmp_path):
    """The buyer is told "48 hours to install". The countdown starts at the
    workbook's Sold Date, so the code dies at the first run at least two days
    later — and even then only if still unclaimed, because a buyer who DID
    activate while the sheet lagged must keep same-device re-activation.
    """
    import datetime

    today = datetime.date(2026, 8, 10)
    fresh = LedgerRow("Sold", datetime.date(2026, 8, 9))
    expired = LedgerRow("Sold", datetime.date(2026, 8, 8))
    undated = LedgerRow("Sold", None)

    assert sold_window_expired(fresh, today, 2) is False
    assert sold_window_expired(expired, today, 2) is True
    assert sold_window_expired(undated, today, 2) is False, "no date, no countdown"

    sql = build_sql(
        {
            "PITW-AAAAA-AAAAA-AAAAA": fresh,
            "PITW-BBBBB-BBBBB-BBBBB": expired,
            "PITW-CCCCC-CCCCC-CCCCC": undated,
        },
        today=today,
    )
    assert (
        "UPDATE codes SET disabled = 0, disabled_reason = NULL "
        "WHERE code_id = 'PITW-AAAAA-AAAAA-AAAAA';" in sql
    ), "inside the window: still activatable"
    assert (
        "WHERE code_id = 'PITW-BBBBB-BBBBB-BBBBB' AND claimed = 0;" in sql
    ), "expired: retired only if never activated"
    assert "activation window expired 2026-08-08 + 2d" in sql
    assert (
        "UPDATE codes SET disabled = 0, disabled_reason = NULL "
        "WHERE code_id = 'PITW-CCCCC-CCCCC-CCCCC';" in sql
    ), "no Sold Date: left active"


def test_sold_date_written_by_hand_or_by_excel_both_count(tmp_path):
    import datetime

    path = _workbook_with_statuses(tmp_path, ["Sold", "Sold", "Sold"])
    workbook = load_workbook(path)
    sheet = workbook["Activation Keys"]
    sheet.cell(row=2, column=4, value=datetime.datetime(2026, 8, 8, 14, 30))  # noqa: DTZ001 - Excel cells are naive
    sheet.cell(row=3, column=4, value="2026-08-08")  # typed ISO
    sheet.cell(row=4, column=4, value="8th of August")  # unreadable
    workbook.save(path)

    rows = read_statuses(path)
    assert rows["PITW-AAAAA-AAAAA-AAAAA"].sold_date == datetime.date(2026, 8, 8)
    assert rows["PITW-BBBBB-BBBBB-BBBBB"].sold_date == datetime.date(2026, 8, 8)
    assert rows["PITW-CCCCC-CCCCC-CCCCC"].sold_date is None


def test_newest_workbook_picks_the_latest_batch(tmp_path):
    for stamp in ("20260101T000000Z", "20260806T072538Z"):
        build_workbook(CODES, tmp_path / f"activation_keys_{stamp}.xlsx", stamp)
    assert newest_workbook(tmp_path).name == "activation_keys_20260806T072538Z.xlsx"
    with pytest.raises(FileNotFoundError):
        newest_workbook(tmp_path / "empty")


def test_worker_refuses_disabled_codes_on_every_path():
    """Source contract: all three Worker handlers check `disabled`.

    The Worker has no JS test harness; this pins the presence of the checks the
    same way the dashboard's JS contracts are pinned. The activation check must
    come before the claimed branch, or a Replaced code's original device could
    still re-activate — the exact hole the sync exists to close.
    """
    worker = (
        Path(__file__).resolve().parents[1]
        / "activation-server" / "src" / "worker.js"
    ).read_text(encoding="utf-8")
    assert worker.count("code_retired") == 3, "activate, download and file must all refuse"
    assert worker.count("disabled FROM codes") == 3, "every SELECT must fetch the flag"
    activate = worker.split("async function handleActivate")[1].split("async function")[0]
    assert activate.index("disabled === 1") < activate.index("claimed === 1")


def test_schema_and_migration_agree_on_the_new_columns():
    server = Path(__file__).resolve().parents[1] / "activation-server"
    schema = (server / "schema.sql").read_text(encoding="utf-8")
    migration = (server / "migrations" / "0001_disabled_codes.sql").read_text(encoding="utf-8")
    for column in ("disabled", "disabled_reason"):
        assert column in schema
        assert column in migration
