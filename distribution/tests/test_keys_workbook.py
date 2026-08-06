"""The activation-key tracker.

LibreOffice is not available on this build machine and the recalc helper needs
POSIX sockets, so the formulas here are checked structurally rather than
evaluated: correct functions, and ranges that match the data exactly. An
off-by-one range would evaluate cleanly and report the wrong count, which is
the failure worth catching.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

DIST = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(DIST.parent))

from distribution.tools.keys_workbook import STATUSES, build_workbook  # noqa: E402

CODES = [f"PITW-AAAAA-BBBBB-{i:05d}" for i in range(1, 8)]

# Excel-2007 era only: these need no _xlfn prefix and LibreOffice evaluates
# them. A post-2007 name written bare becomes a literal #NAME? in the file.
SAFE_FUNCTIONS = {"COUNTA", "COUNTIF"}


@pytest.fixture
def workbook(tmp_path):
    path = build_workbook(CODES, tmp_path / "keys.xlsx", "test batch")
    return load_workbook(path)


def test_it_has_the_three_tabs(workbook):
    assert workbook.sheetnames == ["How to use this", "Activation Keys", "Summary"]


def test_every_code_is_listed_once(workbook):
    sheet = workbook["Activation Keys"]
    listed = [sheet.cell(row=r, column=1).value for r in range(2, len(CODES) + 2)]
    assert listed == CODES
    assert sheet.cell(row=len(CODES) + 2, column=1).value is None


def test_codes_start_unused(workbook):
    sheet = workbook["Activation Keys"]
    assert {sheet.cell(row=r, column=2).value for r in range(2, len(CODES) + 2)} == {"Unused"}


def test_the_status_dropdown_covers_every_row_and_only_valid_states(workbook):
    sheet = workbook["Activation Keys"]
    validations = sheet.data_validations.dataValidation
    assert len(validations) == 1
    validation = validations[0]
    assert str(validation.sqref) == f"B2:B{len(CODES) + 1}"
    assert validation.formula1 == '"' + ",".join(STATUSES) + '"'


def test_summary_formulas_use_only_functions_that_evaluate(workbook):
    for row in workbook["Summary"].iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                for name in set(re.findall(r"([A-Z][A-Z0-9_.]*)\s*\(", cell.value)):
                    assert name in SAFE_FUNCTIONS, f"{cell.coordinate} uses {name}"


def test_summary_ranges_match_the_data_exactly(workbook):
    last = len(CODES) + 1
    for row in workbook["Summary"].iter_rows():
        for cell in row:
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                continue
            for lo, hi in re.findall(r"\$[AB]\$(\d+):\$[AB]\$(\d+)", cell.value):
                assert (int(lo), int(hi)) == (2, last), (
                    f"{cell.coordinate} covers {lo}:{hi}, data is 2:{last}"
                )


def test_the_sheet_name_with_a_space_is_quoted(workbook):
    # Unquoted, a cross-sheet reference containing a space evaluates to #VALUE!.
    for row in workbook["Summary"].iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "Activation Keys" in cell.value:
                assert "'Activation Keys'!" in cell.value, cell.coordinate


def test_revenue_reads_from_the_price_cell_not_a_hardcoded_twenty(workbook):
    summary = workbook["Summary"]
    assert summary["B3"].value == 20
    revenue = next(
        cell.value
        for row in summary.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=($B$3)")
    )
    # Changing the price must change the total; a literal 20 would not.
    assert "$B$3" in revenue
    assert "Sold" in revenue and "Activated" in revenue
    assert "Replaced" not in revenue, "a replacement code is not a second sale"


def test_the_instructions_say_the_file_is_private(workbook):
    text = "\n".join(
        str(cell.value)
        for row in workbook["How to use this"].iter_rows()
        for cell in row
        if cell.value
    )
    assert "KEEP THIS FILE PRIVATE" in text
    assert "one computer only" in text
    assert "gitignored" in text


def test_an_empty_batch_does_not_produce_a_broken_sheet(tmp_path):
    path = build_workbook([], tmp_path / "empty.xlsx", "empty batch")
    sheet = load_workbook(path)["Activation Keys"]
    assert sheet.cell(row=2, column=1).value is None
